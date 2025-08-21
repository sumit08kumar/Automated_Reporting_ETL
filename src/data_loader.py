"""
Data Loading Module for Automated Reporting ETL Pipeline

This module handles the loading of processed data into various output formats
including Excel files with styling, CSV files, and basic visualizations.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import logging
from typing import Dict, List, Optional, Union, Any
from pathlib import Path
from datetime import datetime
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


class DataLoader:
    """
    A class to handle data loading and report generation in various formats.
    """
    
    def __init__(self, output_directory: str, log_level: str = "INFO"):
        """
        Initialize the DataLoader.
        
        Args:
            output_directory (str): Path to the directory for output files
            log_level (str): Logging level (DEBUG, INFO, WARNING, ERROR)
        """
        self.output_directory = Path(output_directory)
        self.output_directory.mkdir(parents=True, exist_ok=True)
        
        # Set up logging
        logging.basicConfig(
            level=getattr(logging, log_level.upper()),
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
        
        # Default styling configuration
        self.default_styles = {
            'header_font': Font(bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
    
    def save_to_csv(self, 
                   df: pd.DataFrame, 
                   filename: str, 
                   **kwargs) -> str:
        """
        Save DataFrame to CSV file.
        
        Args:
            df (pd.DataFrame): DataFrame to save
            filename (str): Name of the output file
            **kwargs: Additional arguments for pandas to_csv()
        
        Returns:
            str: Path to the saved file
        """
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        output_path = self.output_directory / filename
        
        # Default CSV parameters
        csv_params = {
            'index': False,
            'encoding': 'utf-8'
        }
        csv_params.update(kwargs)
        
        try:
            df.to_csv(output_path, **csv_params)
            self.logger.info(f"Successfully saved CSV file: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error saving CSV file {filename}: {str(e)}")
            raise
    
    def save_to_excel_simple(self, 
                           df: pd.DataFrame, 
                           filename: str, 
                           sheet_name: str = 'Data',
                           **kwargs) -> str:
        """
        Save DataFrame to Excel file using pandas (simple version).
        
        Args:
            df (pd.DataFrame): DataFrame to save
            filename (str): Name of the output file
            sheet_name (str): Name of the Excel sheet
            **kwargs: Additional arguments for pandas to_excel()
        
        Returns:
            str: Path to the saved file
        """
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        output_path = self.output_directory / filename
        
        # Default Excel parameters
        excel_params = {
            'index': False,
            'engine': 'openpyxl'
        }
        excel_params.update(kwargs)
        
        try:
            df.to_excel(output_path, sheet_name=sheet_name, **excel_params)
            self.logger.info(f"Successfully saved Excel file: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error saving Excel file {filename}: {str(e)}")
            raise
    
    def save_to_excel_styled(self, 
                           df: pd.DataFrame, 
                           filename: str, 
                           sheet_name: str = 'Data',
                           apply_styling: bool = True,
                           create_table: bool = True) -> str:
        """
        Save DataFrame to Excel file with advanced styling using openpyxl.
        
        Args:
            df (pd.DataFrame): DataFrame to save
            filename (str): Name of the output file
            sheet_name (str): Name of the Excel sheet
            apply_styling (bool): Whether to apply styling
            create_table (bool): Whether to create an Excel table
        
        Returns:
            str: Path to the saved file
        """
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        output_path = self.output_directory / filename
        
        try:
            # First save with pandas
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            if apply_styling:
                # Load the workbook for styling
                wb = load_workbook(output_path)
                ws = wb[sheet_name]
                
                # Apply header styling
                for cell in ws[1]:  # First row (headers)
                    cell.font = self.default_styles['header_font']
                    cell.fill = self.default_styles['header_fill']
                    cell.alignment = self.default_styles['alignment']
                    cell.border = self.default_styles['border']
                
                # Apply border to all cells
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                      min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = self.default_styles['border']
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Create table if requested
                if create_table:
                    from openpyxl.worksheet.table import Table, TableStyleInfo
                    
                    table_range = f"A1:{ws.max_column_letter}{ws.max_row}"
                    table = Table(displayName=f"Table_{sheet_name}", ref=table_range)
                    
                    style = TableStyleInfo(
                        name="TableStyleMedium9", 
                        showFirstColumn=False,
                        showLastColumn=False, 
                        showRowStripes=True, 
                        showColumnStripes=True
                    )
                    table.tableStyleInfo = style
                    ws.add_table(table)
                
                wb.save(output_path)
            
            self.logger.info(f"Successfully saved styled Excel file: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error saving styled Excel file {filename}: {str(e)}")
            raise
    
    def create_summary_sheet(self, 
                           df: pd.DataFrame, 
                           filename: str,
                           summary_config: Dict = None) -> str:
        """
        Create an Excel file with a summary sheet containing KPIs and basic charts.
        
        Args:
            df (pd.DataFrame): DataFrame to analyze
            filename (str): Name of the output file
            summary_config (Dict): Configuration for summary generation
        
        Returns:
            str: Path to the saved file
        """
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        output_path = self.output_directory / filename
        
        if summary_config is None:
            summary_config = {}
        
        try:
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Save original data
                df.to_excel(writer, sheet_name='Data', index=False)
                
                # Create summary data
                summary_data = self._generate_summary_data(df, summary_config)
                summary_df = pd.DataFrame(list(summary_data.items()), 
                                        columns=['Metric', 'Value'])
                
                # Save summary sheet
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Create aggregated data if specified
                if 'group_by' in summary_config:
                    group_cols = summary_config['group_by']
                    agg_funcs = summary_config.get('aggregations', {'count': 'count'})
                    
                    if all(col in df.columns for col in group_cols):
                        agg_df = df.groupby(group_cols).agg(agg_funcs).reset_index()
                        agg_df.to_excel(writer, sheet_name='Aggregated', index=False)
            
            # Add styling and charts
            self._add_summary_styling_and_charts(output_path, df, summary_config)
            
            self.logger.info(f"Successfully created summary report: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error creating summary sheet {filename}: {str(e)}")
            raise
    
    def _generate_summary_data(self, df: pd.DataFrame, config: Dict) -> Dict[str, Any]:
        """
        Generate summary statistics from DataFrame.
        
        Args:
            df (pd.DataFrame): Input DataFrame
            config (Dict): Configuration for summary generation
        
        Returns:
            Dict: Summary statistics
        """
        summary = {}
        
        # Basic statistics
        summary['Total Rows'] = len(df)
        summary['Total Columns'] = len(df.columns)
        summary['Missing Values'] = df.isnull().sum().sum()
        
        # Numeric column statistics
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            for col in numeric_cols:
                summary[f'{col} - Total'] = df[col].sum()
                summary[f'{col} - Average'] = round(df[col].mean(), 2)
                summary[f'{col} - Max'] = df[col].max()
                summary[f'{col} - Min'] = df[col].min()
        
        # Categorical column statistics
        categorical_cols = df.select_dtypes(include=['object', 'category']).columns
        for col in categorical_cols:
            if col != 'source_file' and col != 'source_path':  # Skip metadata columns
                unique_count = df[col].nunique()
                summary[f'{col} - Unique Values'] = unique_count
                if unique_count <= 10:  # Show top values for small categories
                    top_value = df[col].value_counts().index[0]
                    summary[f'{col} - Top Value'] = top_value
        
        # Custom KPIs from config
        custom_kpis = config.get('custom_kpis', {})
        for kpi_name, kpi_formula in custom_kpis.items():
            try:
                result = df.eval(kpi_formula).sum() if 'sum' in kpi_formula else df.eval(kpi_formula).iloc[0]
                summary[kpi_name] = result
            except Exception as e:
                self.logger.warning(f"Failed to calculate custom KPI {kpi_name}: {str(e)}")
        
        return summary
    
    def _add_summary_styling_and_charts(self, 
                                      file_path: str, 
                                      df: pd.DataFrame, 
                                      config: Dict):
        """
        Add styling and charts to the summary Excel file.
        
        Args:
            file_path (str): Path to the Excel file
            df (pd.DataFrame): Original DataFrame
            config (Dict): Configuration for charts and styling
        """
        try:
            wb = load_workbook(file_path)
            
            # Style summary sheet
            if 'Summary' in wb.sheetnames:
                ws_summary = wb['Summary']
                
                # Apply header styling
                for cell in ws_summary[1]:
                    cell.font = self.default_styles['header_font']
                    cell.fill = self.default_styles['header_fill']
                    cell.alignment = self.default_styles['alignment']
                
                # Auto-adjust column widths
                for column in ws_summary.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws_summary.column_dimensions[column_letter].width = adjusted_width
            
            # Add charts if data sheet exists
            if 'Data' in wb.sheetnames:
                ws_data = wb['Data']
                
                # Create a simple bar chart for numeric data
                numeric_cols = df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0 and len(df) > 1:
                    self._add_bar_chart(wb, ws_data, df, numeric_cols[:3])  # Limit to first 3 numeric columns
            
            wb.save(file_path)
            
        except Exception as e:
            self.logger.warning(f"Failed to add styling and charts: {str(e)}")
    
    def _add_bar_chart(self, workbook, worksheet, df: pd.DataFrame, columns: List[str]):
        """
        Add a bar chart to the worksheet.
        
        Args:
            workbook: Excel workbook object
            worksheet: Excel worksheet object
            df (pd.DataFrame): Data for the chart
            columns (List[str]): Columns to include in the chart
        """
        try:
            # Create a new sheet for charts
            if 'Charts' not in workbook.sheetnames:
                chart_sheet = workbook.create_sheet('Charts')
            else:
                chart_sheet = workbook['Charts']
            
            # Prepare data for chart (aggregate if too many rows)
            if len(df) > 20:
                # If too many rows, create aggregated data
                chart_data = df[columns].head(20)
            else:
                chart_data = df[columns]
            
            # Write chart data to chart sheet
            for r_idx, row in enumerate(dataframe_to_rows(chart_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    chart_sheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Create bar chart
            chart = BarChart()
            chart.title = "Data Overview"
            chart.x_axis.title = "Categories"
            chart.y_axis.title = "Values"
            
            # Set data range
            data_range = Reference(chart_sheet, 
                                 min_col=2, max_col=len(columns)+1,
                                 min_row=1, max_row=min(len(chart_data)+1, 21))
            categories = Reference(chart_sheet, 
                                 min_col=1, max_col=1,
                                 min_row=2, max_row=min(len(chart_data)+1, 21))
            
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(categories)
            
            # Add chart to sheet
            chart_sheet.add_chart(chart, "E2")
            
        except Exception as e:
            self.logger.warning(f"Failed to add bar chart: {str(e)}")
    
    def create_visualization(self, 
                           df: pd.DataFrame, 
                           chart_type: str, 
                           filename: str,
                           **kwargs) -> str:
        """
        Create standalone visualizations using matplotlib/seaborn.
        
        Args:
            df (pd.DataFrame): DataFrame to visualize
            chart_type (str): Type of chart ('bar', 'line', 'pie', 'scatter', 'heatmap')
            filename (str): Name of the output file
            **kwargs: Additional arguments for the specific chart type
        
        Returns:
            str: Path to the saved visualization
        """
        if not filename.endswith(('.png', '.jpg', '.pdf')):
            filename += '.png'
        
        output_path = self.output_directory / filename
        
        try:
            plt.figure(figsize=kwargs.get('figsize', (10, 6)))
            
            if chart_type == 'bar':
                x_col = kwargs.get('x', df.columns[0])
                y_col = kwargs.get('y', df.columns[1] if len(df.columns) > 1 else df.columns[0])
                plt.bar(df[x_col], df[y_col])
                plt.xlabel(x_col)
                plt.ylabel(y_col)
                plt.title(kwargs.get('title', f'{y_col} by {x_col}'))
                plt.xticks(rotation=45)
                
            elif chart_type == 'line':
                x_col = kwargs.get('x', df.columns[0])
                y_col = kwargs.get('y', df.columns[1] if len(df.columns) > 1 else df.columns[0])
                plt.plot(df[x_col], df[y_col], marker='o')
                plt.xlabel(x_col)
                plt.ylabel(y_col)
                plt.title(kwargs.get('title', f'{y_col} over {x_col}'))
                plt.xticks(rotation=45)
                
            elif chart_type == 'pie':
                values_col = kwargs.get('values', df.columns[0])
                labels_col = kwargs.get('labels', df.columns[1] if len(df.columns) > 1 else None)
                
                if labels_col:
                    plt.pie(df[values_col], labels=df[labels_col], autopct='%1.1f%%')
                else:
                    plt.pie(df[values_col], autopct='%1.1f%%')
                plt.title(kwargs.get('title', f'Distribution of {values_col}'))
                
            elif chart_type == 'scatter':
                x_col = kwargs.get('x', df.columns[0])
                y_col = kwargs.get('y', df.columns[1] if len(df.columns) > 1 else df.columns[0])
                plt.scatter(df[x_col], df[y_col], alpha=0.6)
                plt.xlabel(x_col)
                plt.ylabel(y_col)
                plt.title(kwargs.get('title', f'{y_col} vs {x_col}'))
                
            elif chart_type == 'heatmap':
                numeric_df = df.select_dtypes(include=[np.number])
                if len(numeric_df.columns) > 1:
                    correlation_matrix = numeric_df.corr()
                    sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', center=0)
                    plt.title(kwargs.get('title', 'Correlation Heatmap'))
                else:
                    raise ValueError("Heatmap requires at least 2 numeric columns")
            
            plt.tight_layout()
            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            plt.close()
            
            self.logger.info(f"Successfully created visualization: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error creating visualization {filename}: {str(e)}")
            plt.close()  # Ensure plot is closed even on error
            raise
    
    def create_multi_format_report(self, 
                                 df: pd.DataFrame, 
                                 base_filename: str,
                                 formats: List[str] = None,
                                 config: Dict = None) -> Dict[str, str]:
        """
        Create reports in multiple formats.
        
        Args:
            df (pd.DataFrame): DataFrame to export
            base_filename (str): Base name for output files
            formats (List[str]): List of formats ('csv', 'excel', 'excel_styled', 'summary')
            config (Dict): Configuration for report generation
        
        Returns:
            Dict[str, str]: Dictionary mapping format names to file paths
        """
        if formats is None:
            formats = ['csv', 'excel_styled', 'summary']
        
        if config is None:
            config = {}
        
        output_files = {}
        
        for format_type in formats:
            try:
                if format_type == 'csv':
                    path = self.save_to_csv(df, f"{base_filename}.csv")
                    output_files['csv'] = path
                    
                elif format_type == 'excel':
                    path = self.save_to_excel_simple(df, f"{base_filename}.xlsx")
                    output_files['excel'] = path
                    
                elif format_type == 'excel_styled':
                    path = self.save_to_excel_styled(df, f"{base_filename}_styled.xlsx")
                    output_files['excel_styled'] = path
                    
                elif format_type == 'summary':
                    path = self.create_summary_sheet(df, f"{base_filename}_summary.xlsx", config)
                    output_files['summary'] = path
                    
            except Exception as e:
                self.logger.error(f"Failed to create {format_type} format: {str(e)}")
        
        self.logger.info(f"Created {len(output_files)} report formats")
        return output_files


# Example usage and testing functions
def main():
    """
    Example usage of the DataLoader class.
    """
    # Create sample data
    sample_data = pd.DataFrame({
        'product': ['Product A', 'Product B', 'Product C', 'Product A', 'Product B'],
        'sales': [1000, 1500, 800, 1200, 1800],
        'quantity': [10, 15, 8, 12, 18],
        'region': ['North', 'South', 'East', 'North', 'South'],
        'date': pd.date_range('2024-01-01', periods=5)
    })
    
    print("Sample data:")
    print(sample_data)
    
    # Initialize loader
    loader = DataLoader("../data/output")
    
    # Create reports in multiple formats
    output_files = loader.create_multi_format_report(
        sample_data, 
        "sample_report",
        formats=['csv', 'excel_styled', 'summary'],
        config={
            'group_by': ['product'],
            'aggregations': {'sales': 'sum', 'quantity': 'sum'},
            'custom_kpis': {
                'Total Revenue': 'sales.sum()',
                'Average Sale': 'sales.mean()'
            }
        }
    )
    
    print("\nGenerated files:")
    for format_name, file_path in output_files.items():
        print(f"  {format_name}: {file_path}")
    
    # Create a visualization
    viz_path = loader.create_visualization(
        sample_data, 
        'bar', 
        'sales_by_product',
        x='product', 
        y='sales',
        title='Sales by Product'
    )
    print(f"  visualization: {viz_path}")


if __name__ == "__main__":
    main()

